#!/usr/bin/env python3
"""
Package flattened rubric PDFs into a Blackboard Ultra Marks and Feedback upload ZIP.

Creates the required structure:
    upload.zip
    ├── AssignmentName.xlsx          ← grade sheet (template with usernames pre-filled)
    ├── u12345678/                   ← one folder per student
    │   └── rubric.pdf               ← flattened rubric
    ├── u23456789/
    │   └── rubric.pdf
    └── ...

Usage:
    python package_for_upload.py <pdf_dir> <assignment_name> <course_id> [output_zip]

Arguments:
    pdf_dir           Directory containing flattened feedback_uXXXXXXXX.pdf files
    assignment_name   Name of the assignment in Blackboard (e.g. "MKM411 Semester Test 1")
    course_id         Blackboard course ID (e.g. "p_ultra_sandbox277_2025")
    output_zip        Optional output ZIP path (default: upload.zip)

Example:
    python package_for_upload.py rubrics/flattened/ "MKM411 Semester Test 1" "MKM411_2026" MKM411_ST1_upload.zip
"""

import sys
import re
import zipfile
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# Extract username from feedback_uXXXXXXXX.pdf
USERNAME_PATTERN = re.compile(r'^feedback_(u\d+)\.pdf$', re.IGNORECASE)


def extract_username(filename: str) -> str | None:
    match = USERNAME_PATTERN.match(filename)
    if match:
        return match.group(1)
    return None


def create_gradebook_excel(output_path: Path, assignment_name: str,
                           course_id: str, usernames: list[str]):
    """Create the Blackboard-format Excel gradebook template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluation Sheet"

    # Row 1: Assignment name
    ws.cell(1, 1, f"Assignment: {assignment_name}")

    # Row 2: Course ID
    ws.cell(2, 1, f"Course ID:{course_id}")

    # Row 3: Headers
    headers = ["USERNAME", "FIRSTNAME", "LASTNAME", "GROUP",
               "SUBMISSION DATE", "Submission Status", "Grade", "Feedback"]
    for col, header in enumerate(headers, 1):
        ws.cell(3, col, header)

    # Row 4+: One row per student, username pre-filled
    for i, username in enumerate(sorted(usernames), 4):
        ws.cell(i, 1, username)                    # USERNAME
        ws.cell(i, 2, "")                          # FIRSTNAME - fill in
        ws.cell(i, 3, "")                          # LASTNAME - fill in
        ws.cell(i, 4, "")                          # GROUP
        ws.cell(i, 5, "")                          # SUBMISSION DATE
        ws.cell(i, 6, "NeedsGrading")              # Submission Status
        ws.cell(i, 7, "!")                          # Grade - fill in
        ws.cell(i, 8, "")                          # Feedback - optional text

    # Auto-fit column widths roughly
    col_widths = [25, 15, 15, 10, 22, 18, 10, 30]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    wb.save(output_path)


def main():
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)

    pdf_dir = Path(sys.argv[1])
    assignment_name = sys.argv[2]
    course_id = sys.argv[3]
    output_zip = Path(sys.argv[4]) if len(sys.argv) >= 5 else Path("upload.zip")

    if not pdf_dir.is_dir():
        print(f"Error: '{pdf_dir}' is not a directory.")
        sys.exit(1)
    if " " in output_zip.name:
        print(f"Error: ZIP filename must not contain spaces. Got '{output_zip.name}'.")
        sys.exit(1)

    # Discover feedback PDFs and extract usernames
    pdf_files = sorted(pdf_dir.glob("feedback_*.pdf"))
    if not pdf_files:
        print(f"No feedback_*.pdf files found in '{pdf_dir}'.")
        sys.exit(1)

    user_pdf_map = {}
    for pdf_path in pdf_files:
        username = extract_username(pdf_path.name)
        if username:
            user_pdf_map[username] = pdf_path
        else:
            print(f"  SKIP: {pdf_path.name} (could not extract username)")

    if not user_pdf_map:
        print("No valid feedback PDFs found.")
        sys.exit(1)

    print(f"Assignment: {assignment_name}")
    print(f"Course ID:  {course_id}")
    print(f"Students:   {len(user_pdf_map)}")
    print(f"Output ZIP: {output_zip}\n")

    # Create the Excel gradebook template
    excel_name = f"{assignment_name.replace(' ', '_')}.xlsx"
    excel_path = pdf_dir / excel_name
    create_gradebook_excel(excel_path, assignment_name, course_id,
                           list(user_pdf_map.keys()))
    print(f"  Created: {excel_name}")

    # Build the ZIP
    with zipfile.ZipFile(output_zip, "w", zipfile.ZIP_DEFLATED) as zf:
        # Add Excel at root level
        zf.write(excel_path, excel_name)

        # Add each rubric in its student folder
        for username, pdf_path in sorted(user_pdf_map.items()):
            arcname = f"{username}/{pdf_path.name}"
            zf.write(pdf_path, arcname)
            print(f"  + {arcname}")

    # Clean up temp Excel
    excel_path.unlink()

    size_mb = output_zip.stat().st_size / (1024 * 1024)
    print(f"\nCreated: {output_zip} ({size_mb:.1f} MB)")

    if size_mb > 250:
        print("WARNING: ZIP exceeds 250 MB — you may need to split into batches.")

    print(f"\nNext steps:")
    print(f"  1. Extract the ZIP")
    print(f"  2. Open {excel_name} and fill in FIRSTNAME, LASTNAME, Grade columns")
    print(f"  3. Re-zip everything (keeping the folder structure)")
    print(f"  4. Upload via Content Market -> Marks and Feedback -> Upload")


if __name__ == "__main__":
    main()